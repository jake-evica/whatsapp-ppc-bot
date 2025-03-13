import pandas as pd
from openpyxl import load_workbook
import datetime


class PPCBidService:
    @staticmethod
    def load_excel_data(file_path):
        """Load Excel file and return workbook object."""
        try:
            return load_workbook(file_path)
        except FileNotFoundError:
            print(f"Error: File '{file_path}' not found.")
            exit()

    @staticmethod
    def get_data_frame(file_path, sheet_name):
        """Load a sheet from an Excel file into a Pandas DataFrame."""
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name).fillna(0)
        except ValueError:
            print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
            exit()

    @staticmethod
    def load_asin_data(file_path):
        """Load ASIN to AOV mapping from 'ASIN Data' sheet if available."""
        wb = PPCBidService.load_excel_data(file_path)
        if "ASIN Data" in wb.sheetnames:
            df_asin = PPCBidService.get_data_frame(file_path, "ASIN Data")
            return dict(zip(df_asin.iloc[:, 0], df_asin.iloc[:, 1]))  # ASIN to AOV mapping
        return {}

    @staticmethod
    def load_campaign_data(file_path):
        """Load campaign data from the second sheet in the Excel file."""
        wb = PPCBidService.load_excel_data(file_path)
        sheet_names = wb.sheetnames
        if len(sheet_names) < 2:
            print("Error: Not enough sheets in the workbook.")
            exit()
        return PPCBidService.get_data_frame(file_path, sheet_names[0])

    @staticmethod
    def calculate_metrics(df, asin_dict):
        """Compute required fields such as ACTC, AOV, % of AOV, and RPC."""
        df["ACTC"] = df["Clicks"] / df["Orders"].replace(0, 1)
        df["AOV"] = df.apply(
            lambda row: row["Sales"] / row["Orders"] if row["Orders"] > 0 else asin_dict.get(row["ASIN (Informational only)"], 0), axis=1
        )
        df["% of AOV"] = df["Spend"] / df["AOV"].replace(0, 1)
        df["RPC"] = df["Sales"] / df["Clicks"].replace(0, 1)
        return df

    @staticmethod
    def adjust_bid(row, target_acos, increase_spend):
        """Apply bid adjustments based on ACOS conditions."""
        acos = row["ACOS"]
        orders = row["Orders"]
        aov_percent = row["% of AOV"]
        impressions = row["Impressions"]

        if acos >= target_acos * 1.1:  # High ACOS (Reduce Bid)
            return row["RPC"] * target_acos
        elif acos <= target_acos * 0.9 and orders > 1:  # Low ACOS (Increase Bid)
            return row["Old Bid"] * 1.15
        elif acos <= target_acos * 0.9 and orders == 1:
            return row["Old Bid"] * 1.05
        elif acos == 0 and aov_percent >= target_acos * 0.9:
            return row["Sales"] / (row["Clicks"] + row["Impressions"]) * target_acos
        elif increase_spend and acos == 0 and aov_percent <= 0.1 and impressions >= 0.003:
            return row["Old Bid"] * 1.05
        return row["Old Bid"]

    @staticmethod
    def optimize_bids(file_path, target_acos = 0.3, increase_spend = True):
        """
        Main function that runs all processes and returns the final DataFrame.
        :param file_path: Path to the Excel file.
        :param target_acos: Target ACOS value (e.g., 0.3 for 30%).
        :param increase_spend: Boolean flag (True/False) to increase spend on low-spend keywords.
        :return: Processed DataFrame with optimized bids.
        """
        # Step 1: Load data
        asin_dict = PPCBidService.load_asin_data(file_path)
        df = PPCBidService.load_campaign_data(file_path)

        # Ensure "Old Bid" column exists
        if "Old Bid" not in df.columns:
            df["Old Bid"] = df["Bid"]

        # Step 2: Process data
        df = PPCBidService.calculate_metrics(df, asin_dict)

        # Step 3: Apply bid adjustments
        df["New Bid"] = df.apply(lambda row: PPCBidService.adjust_bid(row, target_acos, increase_spend), axis=1)
        df["Update"] = df["New Bid"] != df["Old Bid"]

        # Step 4: Filter rows that require an update
        df = df[df["Update"]]
        return df


# Run the script
if __name__ == "__main__":
    file_path = "data/raw/sample_bid_0001.xlsx"
    df_optimized = PPCBidService.optimize_bids(file_path)

    output_file = f"data/processed/Bid_Optimization_{datetime.datetime.now().strftime('%m%d%y')}.xlsx"
    df_optimized.to_excel(output_file, index=False)
