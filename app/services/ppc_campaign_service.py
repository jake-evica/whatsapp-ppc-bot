import pandas as pd
from openpyxl import Workbook
from typing import List, Tuple


class PPCCampaignService:
    @staticmethod
    def load_data(file_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        xls = pd.ExcelFile(file_path)
        df_sponsored = xls.parse('Sponsored Products Campaigns')
        df_search_terms = xls.parse('SP Search Term Report')
        df_asin_list = xls.parse('ASIN List')
        return df_sponsored, df_search_terms, df_asin_list

    @staticmethod
    def filter_keywords(df_search_terms: pd.DataFrame, acos_threshold: float, brand_exclusions: List[str]) -> Tuple[List[Tuple[str, int, float]], List[Tuple[str, int, float]]]:
        search_term_col = "Keyword ID"
        orders_col = "Orders"
        acos_col = "ACOS"

        filtered_terms = df_search_terms[
            (df_search_terms[orders_col] >= 1) &
            (df_search_terms[acos_col] < acos_threshold)
        ].copy()

        regular_keywords = []
        targeting_keywords = []

        brand_exclusions = [" " + brand.lower() + " " for brand in (brand_exclusions or [])]

        for _, row in filtered_terms.iterrows():
            search_term = str(row[search_term_col]).strip().lower()
            orders = row[orders_col]
            acos = row[acos_col]

            if any(brand in f" {search_term} " for brand in brand_exclusions):
                continue

            if search_term.startswith("b0"):  # ASIN targeting
                targeting_keywords.append((search_term, orders, acos))
            else:
                regular_keywords.append((search_term, orders, acos))

        return regular_keywords, targeting_keywords

    @staticmethod
    def generate_output(regular_keywords: List[Tuple[str, int, float]], 
                        targeting_keywords: List[Tuple[str, int, float]], 
                        match_type: str) -> str:
        wb = Workbook()
        sheets = {
            "3+ Orders": wb.active,
            "1-2 Orders": wb.create_sheet("1-2 Orders"),
            "Product Targets": wb.create_sheet("Product Targets"),
        }
        sheets["3+ Orders"].title = "3+ Orders"

        headers = ["Keyword", "Orders", "ACOS", "Match Type"]
        for sheet in sheets.values():
            sheet.append(headers)

        for keyword, orders, acos in regular_keywords:
            target_sheet = sheets["3+ Orders"] if orders >= 3 else sheets["1-2 Orders"]
            target_sheet.append([keyword, orders, acos, match_type])

        for keyword, orders, acos in targeting_keywords:
            sheets["Product Targets"].append([keyword, orders, acos, "ASIN Targeting"])

        output_file = "data/processed/amazon_ppc_campaign.xlsx"
        wb.save(output_file)
        return output_file

    @staticmethod
    def create_campaign(file_path: str, acos_threshold: float = 0.30, match_type: str = "exact", brand_exclusions: List[str] = None) -> Tuple[str, List[Tuple[str, int, float]]]:
        df_sponsored, df_search_terms, df_asin_list = PPCCampaignService.load_data(file_path)
        regular_keywords, targeting_keywords = PPCCampaignService.filter_keywords(df_search_terms, acos_threshold, brand_exclusions)
        output_file = PPCCampaignService.generate_output(regular_keywords, targeting_keywords, match_type)
        return output_file, regular_keywords